from pathlib import Path
import re
import pandas as pd


# =========================================================
# PATHS
# =========================================================

BASE_DIR = Path(__file__).resolve().parent
INPUT_DIR = BASE_DIR / "inputs"
OUTPUT_DIR = BASE_DIR / "outputs"

ATT_FILE = INPUT_DIR / "att.xlsx"
INTUNE_FILE = INPUT_DIR / "Intune.csv"
EXCLUSIONS_FILE = INPUT_DIR / "imei_exclusions.csv"
DISABLED_USERS_FILE = INPUT_DIR / "disabled_ad_users.csv"

OUTPUT_DIR.mkdir(exist_ok=True)

MASTER_WORKBOOK = "imei_audit_master.xlsx"
REVIEW_WORKBOOK = "imei_review_workbook.xlsx"
DEPT_WORKBOOK = "imei_department_views.xlsx"


# =========================================================
# HELPERS
# =========================================================

def clean_column_name(col):
    if pd.isna(col):
        return ""
    return " ".join(str(col).replace("\n", " ").split()).strip()


def normalize_imei(value):
    if pd.isna(value):
        return ""

    text = str(value).strip()
    if text.endswith(".0"):
        text = text[:-2]

    digits = "".join(ch for ch in text if ch.isdigit())
    return digits if len(digits) == 15 else ""


def normalize_name(value):
    if pd.isna(value):
        return ""
    text = str(value).strip().lower()
    text = re.sub(r"\s+", " ", text)
    return text


def safe_filename(name):
    if name is None:
        return "UNKNOWN"
    text = str(name).strip()
    if not text:
        return "UNKNOWN"
    text = re.sub(r'[<>:"/\\\\|?*]', "_", text)
    text = re.sub(r"\s+", "_", text)
    return text[:80]


def autosize(ws):
    for column in ws.columns:
        max_len = 0
        col_letter = column[0].column_letter
        for cell in column:
            try:
                cell_val = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(cell_val))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, 45)


def delete_if_exists(path: Path):
    try:
        if path.exists():
            path.unlink()
    except PermissionError:
        raise PermissionError(
            f"Cannot overwrite '{path}'. Close it in Excel and run again."
        )


def cleanup_outputs():
    for filename in [
        "imei_audit_full.csv",
        "imei_missing_from_intune_raw.csv",
        "imei_missing_from_intune_cleaned.csv",
        "imei_excluded_from_review.csv",
        "imei_disabled_ad_user_matches.csv",
        "imei_action_list.csv",
        "imei_matched_in_intune.csv",
        "intune_devices_without_imei.csv",
        "imei_department_summary.csv",
        "imei_department_summary_raw.csv",
        "imei_department_summary_cleaned.csv",
        "imei_summary.csv",
        MASTER_WORKBOOK,
        REVIEW_WORKBOOK,
        DEPT_WORKBOOK,
    ]:
        delete_if_exists(OUTPUT_DIR / filename)


# =========================================================
# INPUT LOADERS
# =========================================================

def detect_att_header(att_file):
    header_candidates = [5, 7]

    phone_cols = ["Wireless number", "Wireless Number"]
    name_cols = ["Wireless user name", "Wireless User Full Name"]
    imei_cols = ["Device IMEI", "Phone or Device ID (IMEI)"]
    dept_cols = ["DEPARTMENT", "Department", "department"]

    for header in header_candidates:
        df = pd.read_excel(att_file, header=header, dtype=str)
        df.columns = [clean_column_name(c) for c in df.columns]

        phone = next((c for c in phone_cols if c in df.columns), None)
        name = next((c for c in name_cols if c in df.columns), None)
        imei = next((c for c in imei_cols if c in df.columns), None)
        dept = next((c for c in dept_cols if c in df.columns), None)

        if phone and name and imei:
            return df, phone, name, imei, dept

    raise ValueError("Could not detect AT&T report format.")


def load_exclusions(exclusions_file: Path):
    empty_df = pd.DataFrame(columns=[
        "AT&T IMEI",
        "Department",
        "exclude_reason",
        "exclude_scope",
        "notes",
    ])

    if not exclusions_file.exists():
        return empty_df

    try:
        exclusions = pd.read_csv(exclusions_file, dtype=str).fillna("")
    except pd.errors.EmptyDataError:
        return empty_df

    exclusions.columns = [clean_column_name(c) for c in exclusions.columns]

    if "AT&T IMEI" not in exclusions.columns:
        raise ValueError(
            f"Exclusions file missing required column 'AT&T IMEI'. Found: {list(exclusions.columns)}"
        )

    for col in ["Department", "exclude_reason", "exclude_scope", "notes"]:
        if col not in exclusions.columns:
            exclusions[col] = ""

    exclusions["AT&T IMEI"] = exclusions["AT&T IMEI"].apply(normalize_imei)
    exclusions = exclusions[exclusions["AT&T IMEI"] != ""].copy()
    exclusions = exclusions.drop_duplicates(subset=["AT&T IMEI"], keep="first")

    return exclusions[["AT&T IMEI", "Department", "exclude_reason", "exclude_scope", "notes"]]


def load_disabled_users(disabled_users_file: Path):
    empty_df = pd.DataFrame(columns=[
        "DisplayName",
        "SamAccountName",
        "UserPrincipalName",
        "Department",
        "display_norm",
    ])

    if not disabled_users_file.exists():
        return empty_df

    try:
        disabled = pd.read_csv(disabled_users_file, dtype=str).fillna("")
    except pd.errors.EmptyDataError:
        return empty_df

    disabled.columns = [clean_column_name(c) for c in disabled.columns]

    if "DisplayName" not in disabled.columns:
        raise ValueError(
            f"Disabled users file missing required column 'DisplayName'. Found: {list(disabled.columns)}"
        )

    for col in ["SamAccountName", "UserPrincipalName", "Department"]:
        if col not in disabled.columns:
            disabled[col] = ""

    disabled["display_norm"] = disabled["DisplayName"].apply(normalize_name)
    disabled = disabled[disabled["display_norm"] != ""].copy()
    disabled = disabled.drop_duplicates(subset=["display_norm"], keep="first")

    return disabled[["DisplayName", "SamAccountName", "UserPrincipalName", "Department", "display_norm"]]


# =========================================================
# EXCLUSIONS / TAGGING
# =========================================================

def apply_exclusions(missing_df: pd.DataFrame, exclusions_df: pd.DataFrame):
    missing_df = missing_df.copy()

    if missing_df.empty:
        cleaned = missing_df.copy()
        excluded = pd.DataFrame(columns=list(missing_df.columns) + [
            "exclude_reason", "exclude_scope", "notes", "Review Status"
        ])
        return cleaned, excluded

    if exclusions_df.empty:
        cleaned = missing_df.copy()
        cleaned["Review Status"] = "Needs Review"
        excluded = pd.DataFrame(columns=list(missing_df.columns) + [
            "exclude_reason", "exclude_scope", "notes", "Review Status"
        ])
        return cleaned, excluded

    excluded = missing_df.merge(
        exclusions_df,
        on="AT&T IMEI",
        how="inner",
        suffixes=("", "_excl"),
    ).copy()
    excluded["Review Status"] = "Excluded"

    cleaned = missing_df[
        ~missing_df["AT&T IMEI"].isin(exclusions_df["AT&T IMEI"])
    ].copy()
    cleaned["Review Status"] = "Needs Review"

    return cleaned, excluded


def apply_pattern_exclusions(missing_df: pd.DataFrame):
    df = missing_df.copy()

    extra_cols = ["exclude_reason", "exclude_scope", "notes", "Review Status"]
    base_cols = list(df.columns)
    final_cols = base_cols + [col for col in extra_cols if col not in base_cols]

    if df.empty:
        return df, pd.DataFrame(columns=final_cols)

    if "AT&T User" not in df.columns:
        return df, pd.DataFrame(columns=final_cols)

    att_user_upper = df["AT&T User"].fillna("").astype(str).str.upper()
    dept_upper = (
        df["Department"].fillna("").astype(str).str.upper()
        if "Department" in df.columns
        else pd.Series("", index=df.index)
    )

    pattern_rules = [
        ("Meraki firewall device", att_user_upper.str.contains(r"\bMERAKI\b", na=False)),
        ("Desk phone", att_user_upper.str.contains(r"DESK PHONE", na=False)),
        ("Conference room device", att_user_upper.str.contains(r"CONFERENCE ROOM", na=False)),
        ("Fax line/device", att_user_upper.str.contains(r"\bFAX\b", na=False)),
        ("Modem device", att_user_upper.str.contains(r"\bMODEM\b", na=False)),
        ("Hotspot device", att_user_upper.str.contains(r"\bHOTSPOT\b", na=False)),
        ("Router device", att_user_upper.str.contains(r"\bROUTER\b", na=False)),
        ("Cradlepoint device", att_user_upper.str.contains(r"CRADLEPOINT", na=False)),
        ("Firewall department device", dept_upper.str.contains(r"FIREWALL", na=False)),
    ]

    exclude_mask = pd.Series(False, index=df.index)
    exclude_reason = pd.Series("", index=df.index)

    for reason, rule_mask in pattern_rules:
        new_matches = rule_mask & ~exclude_mask
        exclude_reason.loc[new_matches] = reason
        exclude_mask = exclude_mask | rule_mask

    excluded = df[exclude_mask].copy()
    cleaned = df[~exclude_mask].copy()

    if not excluded.empty:
        excluded["exclude_reason"] = exclude_reason.loc[excluded.index]
        excluded["exclude_scope"] = "pattern"
        excluded["notes"] = "Auto-excluded by pattern rule"
        excluded["Review Status"] = "Excluded"

    if not cleaned.empty:
        cleaned["Review Status"] = "Needs Review"

    excluded = excluded.reindex(columns=final_cols)

    return cleaned, excluded


def tag_disabled_ad_users(df: pd.DataFrame, disabled_users_df: pd.DataFrame):
    df = df.copy()

    if df.empty:
        for col in [
            "AD Disabled User",
            "AD Disabled DisplayName",
            "AD Disabled UPN",
            "AD Disabled Department",
            "AD Disabled SamAccountName",
        ]:
            df[col] = ""
        return df, df.copy()

    df["att_user_norm"] = df["AT&T User"].apply(normalize_name)

    if disabled_users_df.empty:
        df["AD Disabled User"] = "No"
        df["AD Disabled DisplayName"] = ""
        df["AD Disabled UPN"] = ""
        df["AD Disabled Department"] = ""
        df["AD Disabled SamAccountName"] = ""
        df = df.drop(columns=["att_user_norm"], errors="ignore")
        return df, pd.DataFrame(columns=df.columns)

    disabled_users_df = disabled_users_df[disabled_users_df["display_norm"] != ""].copy()

    named_rows = df[df["att_user_norm"] != ""].copy()
    blank_rows = df[df["att_user_norm"] == ""].copy()

    tagged_named = named_rows.merge(
        disabled_users_df,
        left_on="att_user_norm",
        right_on="display_norm",
        how="left",
        suffixes=("", "_ad"),
    )

    tagged_named["AD Disabled User"] = tagged_named["DisplayName"].apply(
        lambda x: "Yes" if pd.notna(x) and str(x).strip() != "" else "No"
    )
    tagged_named["AD Disabled DisplayName"] = tagged_named["DisplayName"].fillna("")
    tagged_named["AD Disabled UPN"] = tagged_named["UserPrincipalName"].fillna("")
    tagged_named["AD Disabled Department"] = tagged_named["Department_ad"].fillna("") if "Department_ad" in tagged_named.columns else ""
    tagged_named["AD Disabled SamAccountName"] = tagged_named["SamAccountName"].fillna("")

    blank_rows["AD Disabled User"] = "No"
    blank_rows["AD Disabled DisplayName"] = ""
    blank_rows["AD Disabled UPN"] = ""
    blank_rows["AD Disabled Department"] = ""
    blank_rows["AD Disabled SamAccountName"] = ""

    tagged = pd.concat([tagged_named, blank_rows], ignore_index=True)
    disabled_matches = tagged[tagged["AD Disabled User"] == "Yes"].copy()

    drop_cols = [
        "att_user_norm",
        "display_norm",
        "DisplayName",
        "SamAccountName",
        "UserPrincipalName",
        "Department_ad",
    ]
    tagged = tagged.drop(columns=[c for c in drop_cols if c in tagged.columns], errors="ignore")
    disabled_matches = disabled_matches.drop(columns=[c for c in drop_cols if c in disabled_matches.columns], errors="ignore")

    return tagged, disabled_matches


def classify_missing_devices(df: pd.DataFrame):
    df = df.copy()

    if df.empty:
        df["Shared/Open Device"] = ""
        df["Likely iPad Device"] = ""
        df["Device Classification"] = ""
        return df

    att_user_upper = df["AT&T User"].fillna("").astype(str).str.upper()
    dept_upper = df["Department"].fillna("").astype(str).str.upper()

    df["Shared/Open Device"] = "No"
    df["Likely iPad Device"] = "No"
    df["Device Classification"] = "Needs Review"

    shared_mask = (
        att_user_upper.str.contains(r"\bOPEN\b", na=False) |
        att_user_upper.str.contains(r"ON CALL", na=False)
    )

    ipad_mask = (
        att_user_upper.str.contains(r"IPAD", na=False) |
        dept_upper.str.contains(r"IPAD", na=False)
    )

    df.loc[shared_mask, "Shared/Open Device"] = "Yes"
    df.loc[ipad_mask, "Likely iPad Device"] = "Yes"
    df.loc[shared_mask, "Device Classification"] = "Shared/Open Device"

    if "AD Disabled User" in df.columns:
        df.loc[df["AD Disabled User"] == "Yes", "Device Classification"] = "Disabled AD User"

    return df


# =========================================================
# SUMMARIES / VALIDATION
# =========================================================

def build_department_summary(report_df: pd.DataFrame):
    dept_summary = (
        report_df.groupby(["Department", "Match Status"], dropna=False)
        .size()
        .unstack(fill_value=0)
        .reset_index()
    )

    if "Matched in Intune" not in dept_summary.columns:
        dept_summary["Matched in Intune"] = 0
    if "Missing from Intune" not in dept_summary.columns:
        dept_summary["Missing from Intune"] = 0

    dept_summary["Total AT&T Devices"] = dept_summary["Matched in Intune"] + dept_summary["Missing from Intune"]

    return dept_summary[
        ["Department", "Total AT&T Devices", "Matched in Intune", "Missing from Intune"]
    ].sort_values(by=["Missing from Intune", "Department"], ascending=[False, True])


def build_cleaned_department_summary(report_df: pd.DataFrame, missing_cleaned_df: pd.DataFrame):
    matched_counts = (
        report_df[report_df["Match Status"] == "Matched in Intune"]
        .groupby("Department", dropna=False)
        .size()
        .rename("Matched in Intune")
    )

    missing_counts = (
        missing_cleaned_df.groupby("Department", dropna=False)
        .size()
        .rename("Missing from Intune")
    )

    dept_summary = pd.concat([matched_counts, missing_counts], axis=1).fillna(0).reset_index()

    if "Matched in Intune" not in dept_summary.columns:
        dept_summary["Matched in Intune"] = 0
    if "Missing from Intune" not in dept_summary.columns:
        dept_summary["Missing from Intune"] = 0

    dept_summary["Matched in Intune"] = dept_summary["Matched in Intune"].astype(int)
    dept_summary["Missing from Intune"] = dept_summary["Missing from Intune"].astype(int)
    dept_summary["Total AT&T Devices"] = dept_summary["Matched in Intune"] + dept_summary["Missing from Intune"]

    return dept_summary[
        ["Department", "Total AT&T Devices", "Matched in Intune", "Missing from Intune"]
    ].sort_values(by=["Missing from Intune", "Department"], ascending=[False, True])


def build_classification_summary(df: pd.DataFrame):
    if df.empty:
        return pd.DataFrame(columns=["Device Classification", "Count"])

    summary = (
        df.groupby("Device Classification", dropna=False)
        .size()
        .reset_index(name="Count")
        .sort_values(by="Count", ascending=False)
    )
    return summary


def build_top_values(df: pd.DataFrame, column_name: str, top_n: int = 50):
    if df.empty or column_name not in df.columns:
        return pd.DataFrame(columns=[column_name, "Count"])

    vc = df[column_name].fillna("").astype(str).value_counts().head(top_n).reset_index()
    vc.columns = [column_name, "Count"]
    return vc


def build_validation_samples(action_list: pd.DataFrame, disabled_action: pd.DataFrame, excluded: pd.DataFrame):
    blocks = []

    if not action_list.empty:
        sample = action_list.head(15).copy()
        sample.insert(0, "Validation Bucket", "Action List Sample")
        blocks.append(sample)

    if not disabled_action.empty:
        sample = disabled_action.head(10).copy()
        sample.insert(0, "Validation Bucket", "Disabled AD Sample")
        blocks.append(sample)

    if not excluded.empty:
        sample = excluded.head(10).copy()
        sample.insert(0, "Validation Bucket", "Excluded Sample")
        blocks.append(sample)

    if blocks:
        return pd.concat(blocks, ignore_index=True)

    return pd.DataFrame(columns=["Validation Bucket"])


# =========================================================
# WORKBOOK WRITERS
# =========================================================

def style_workbook(writer):
    from openpyxl.styles import PatternFill

    green = PatternFill(fill_type="solid", fgColor="C6EFCE")
    red = PatternFill(fill_type="solid", fgColor="F4CCCC")
    yellow = PatternFill(fill_type="solid", fgColor="FFF2CC")
    gray = PatternFill(fill_type="solid", fgColor="D9D9D9")
    blue = PatternFill(fill_type="solid", fgColor="CFE2F3")

    for sheet in writer.sheets.values():
        sheet.freeze_panes = "A2"
        autosize(sheet)

    if "Full Audit" in writer.sheets:
        ws = writer.sheets["Full Audit"]
        headers = [c.value for c in ws[1]]
        if "Match Status" in headers:
            status_col = headers.index("Match Status") + 1
            for r in range(2, ws.max_row + 1):
                status = ws.cell(row=r, column=status_col).value
                fill = green if status == "Matched in Intune" else red
                for c in range(1, ws.max_column + 1):
                    ws.cell(row=r, column=c).fill = fill

    for sheet_name, fill in {
        "Intune No IMEI": yellow,
        "Excluded": gray,
        "Disabled AD Users": blue,
        "Action List": red,
    }.items():
        if sheet_name in writer.sheets:
            ws = writer.sheets[sheet_name]
            for r in range(2, ws.max_row + 1):
                for c in range(1, ws.max_column + 1):
                    ws.cell(row=r, column=c).fill = fill

    if "Missing Cleaned" in writer.sheets:
        ws = writer.sheets["Missing Cleaned"]
        headers = [c.value for c in ws[1]]
        ad_disabled_col = headers.index("AD Disabled User") + 1 if "AD Disabled User" in headers else None
        classification_col = headers.index("Device Classification") + 1 if "Device Classification" in headers else None

        for r in range(2, ws.max_row + 1):
            row_fill = red
            if ad_disabled_col and ws.cell(row=r, column=ad_disabled_col).value == "Yes":
                row_fill = blue
            elif classification_col and ws.cell(row=r, column=classification_col).value == "Shared/Open Device":
                row_fill = yellow

            for c in range(1, ws.max_column + 1):
                ws.cell(row=r, column=c).fill = row_fill


def write_master_workbook(
    report,
    missing_raw,
    missing_tagged,
    matched,
    excluded,
    disabled_action,
    action_list,
    intune_no_imei_report,
    dept_summary_raw,
    dept_summary_cleaned,
    classification_summary,
    top_users,
    top_departments,
    validation_samples,
    summary,
):
    path = OUTPUT_DIR / MASTER_WORKBOOK

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        report.to_excel(writer, sheet_name="Full Audit", index=False)
        missing_raw.to_excel(writer, sheet_name="Missing Raw", index=False)
        missing_tagged.to_excel(writer, sheet_name="Missing Cleaned", index=False)
        matched.to_excel(writer, sheet_name="Matched", index=False)
        excluded.to_excel(writer, sheet_name="Excluded", index=False)
        disabled_action.to_excel(writer, sheet_name="Disabled AD Users", index=False)
        action_list.to_excel(writer, sheet_name="Action List", index=False)
        intune_no_imei_report.to_excel(writer, sheet_name="Intune No IMEI", index=False)
        dept_summary_raw.to_excel(writer, sheet_name="By Department Raw", index=False)
        dept_summary_cleaned.to_excel(writer, sheet_name="By Department Cleaned", index=False)
        classification_summary.to_excel(writer, sheet_name="Classification Summary", index=False)
        top_users.to_excel(writer, sheet_name="Top AT&T Users", index=False)
        top_departments.to_excel(writer, sheet_name="Top Departments", index=False)
        validation_samples.to_excel(writer, sheet_name="Validation Samples", index=False)
        summary.to_excel(writer, sheet_name="Summary", index=False)

        style_workbook(writer)


def write_review_workbook(action_list, disabled_action, excluded, classification_summary):
    path = OUTPUT_DIR / REVIEW_WORKBOOK

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        action_list.to_excel(writer, sheet_name="Action List", index=False)
        disabled_action.to_excel(writer, sheet_name="Disabled AD Users", index=False)
        excluded.to_excel(writer, sheet_name="Excluded", index=False)
        classification_summary.to_excel(writer, sheet_name="Classification Summary", index=False)

        style_workbook(writer)


def write_department_workbook(action_list, disabled_action):
    path = OUTPUT_DIR / DEPT_WORKBOOK

    action_by_dept = (
        action_list.groupby("Department", dropna=False).size().reset_index(name="Action List Count")
        if not action_list.empty else pd.DataFrame(columns=["Department", "Action List Count"])
    )
    disabled_by_dept = (
        disabled_action.groupby("Department", dropna=False).size().reset_index(name="Disabled AD Count")
        if not disabled_action.empty else pd.DataFrame(columns=["Department", "Disabled AD Count"])
    )

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        action_by_dept.to_excel(writer, sheet_name="Action by Department", index=False)
        disabled_by_dept.to_excel(writer, sheet_name="Disabled AD by Department", index=False)

        if not action_list.empty:
            action_list.sort_values(["Department", "AT&T User"]).to_excel(writer, sheet_name="Action Detail", index=False)
        else:
            pd.DataFrame([{"Status": "No action rows"}]).to_excel(writer, sheet_name="Action Detail", index=False)

        if not disabled_action.empty:
            disabled_action.sort_values(["Department", "AT&T User"]).to_excel(writer, sheet_name="Disabled AD Detail", index=False)
        else:
            pd.DataFrame([{"Status": "No disabled AD rows"}]).to_excel(writer, sheet_name="Disabled AD Detail", index=False)

        style_workbook(writer)


# =========================================================
# MAIN
# =========================================================

def main():
    cleanup_outputs()

    print("Loading AT&T report...")
    att_df, phone_col, name_col, imei_col, dept_col = detect_att_header(ATT_FILE)

    print("Loading Intune report...")
    if INTUNE_FILE.suffix.lower() == ".csv":
        intune_df = pd.read_csv(INTUNE_FILE, dtype=str)
    else:
        intune_df = pd.read_excel(INTUNE_FILE, dtype=str)

    intune_df.columns = [clean_column_name(c) for c in intune_df.columns]

    required_intune_cols = ["IMEI", "Device name", "Primary user UPN", "Compliance", "Last check-in"]
    missing_intune_cols = [c for c in required_intune_cols if c not in intune_df.columns]
    if missing_intune_cols:
        raise ValueError(f"Intune file is missing required columns: {missing_intune_cols}")

    print("Loading exclusions...")
    exclusions_df = load_exclusions(EXCLUSIONS_FILE)

    print("Loading disabled AD users...")
    disabled_users_df = load_disabled_users(DISABLED_USERS_FILE)

    print("Normalizing IMEIs...")
    att_df["imei"] = att_df[imei_col].apply(normalize_imei)
    intune_df["imei"] = intune_df["IMEI"].apply(normalize_imei)

    att_valid = att_df[att_df["imei"] != ""].copy()

    intune_with_imei = intune_df[intune_df["imei"] != ""].copy()
    intune_without_imei = intune_df[intune_df["imei"] == ""].copy()
    intune_with_imei = intune_with_imei.drop_duplicates(subset=["imei"])

    merged = att_valid.merge(
        intune_with_imei[["imei", "Device name", "Primary user UPN", "Compliance", "Last check-in"]],
        on="imei",
        how="left",
    )

    merged["Match Status"] = merged["Device name"].apply(
        lambda x: "Matched in Intune" if pd.notna(x) and str(x).strip() else "Missing from Intune"
    )

    report = pd.DataFrame({
        "Department": merged[dept_col].fillna("") if dept_col else "",
        "AT&T User": merged[name_col].fillna(""),
        "AT&T Wireless Number": merged[phone_col].fillna(""),
        "AT&T IMEI": merged["imei"],
        "Match Status": merged["Match Status"],
        "Intune Device Name": merged["Device name"].fillna(""),
        "Intune Primary User": merged["Primary user UPN"].fillna(""),
        "Intune Compliance": merged["Compliance"].fillna(""),
        "Intune Last Check-in": merged["Last check-in"].fillna(""),
    })

    missing_raw = report[report["Match Status"] == "Missing from Intune"].copy()
    matched = report[report["Match Status"] == "Matched in Intune"].copy()

    missing_after_manual, excluded_manual = apply_exclusions(missing_raw, exclusions_df)
    missing_cleaned, excluded_pattern = apply_pattern_exclusions(missing_after_manual)

    excluded_manual = excluded_manual.loc[:, ~excluded_manual.columns.duplicated()].copy()
    excluded_pattern = excluded_pattern.loc[:, ~excluded_pattern.columns.duplicated()].copy()
    excluded = pd.concat([excluded_manual, excluded_pattern], ignore_index=True)

    missing_tagged, disabled_matches = tag_disabled_ad_users(missing_cleaned, disabled_users_df)
    missing_tagged = classify_missing_devices(missing_tagged)

    disabled_action = missing_tagged[missing_tagged["AD Disabled User"] == "Yes"].copy()
    action_list = missing_tagged[missing_tagged["AD Disabled User"] == "No"].copy()

    dept_summary_raw = build_department_summary(report)
    dept_summary_cleaned = build_cleaned_department_summary(report, missing_tagged)
    classification_summary = build_classification_summary(missing_tagged)
    top_users = build_top_values(action_list, "AT&T User", 50)
    top_departments = build_top_values(action_list, "Department", 50)
    validation_samples = build_validation_samples(action_list, disabled_action, excluded)

    intune_no_imei_report = pd.DataFrame({
        "Intune Device Name": intune_without_imei["Device name"].fillna(""),
        "Intune Primary User": intune_without_imei["Primary user UPN"].fillna(""),
        "Intune Compliance": intune_without_imei["Compliance"].fillna(""),
        "Intune Last Check-in": intune_without_imei["Last check-in"].fillna(""),
        "Original IMEI Value": intune_without_imei["IMEI"].fillna(""),
    })

    action_list = action_list.sort_values(by=["Department", "AT&T User", "AT&T Wireless Number"])
    disabled_action = disabled_action.sort_values(by=["Department", "AT&T User", "AT&T Wireless Number"])

    summary = pd.DataFrame([
        {"Metric": "AT&T total rows", "Value": len(att_df)},
        {"Metric": "AT&T valid IMEIs", "Value": len(att_valid)},
        {"Metric": "Intune total rows", "Value": len(intune_df)},
        {"Metric": "Intune with IMEI", "Value": len(intune_with_imei)},
        {"Metric": "Intune without IMEI", "Value": len(intune_without_imei)},
        {"Metric": "Matched", "Value": len(matched)},
        {"Metric": "Missing (raw)", "Value": len(missing_raw)},
        {"Metric": "Excluded from review", "Value": len(excluded)},
        {"Metric": "Missing (cleaned before AD disabled tagging)", "Value": len(missing_cleaned)},
        {"Metric": "Disabled AD user matches", "Value": len(disabled_matches)},
        {"Metric": "Shared/Open Device count", "Value": len(missing_tagged[missing_tagged["Device Classification"] == "Shared/Open Device"])},
        {"Metric": "Likely iPad count", "Value": len(missing_tagged[missing_tagged["Likely iPad Device"] == "Yes"])},
        {"Metric": "Action list count", "Value": len(action_list)},
    ])

    # CSV outputs
    report.to_csv(OUTPUT_DIR / "imei_audit_full.csv", index=False)
    missing_raw.to_csv(OUTPUT_DIR / "imei_missing_from_intune_raw.csv", index=False)
    missing_tagged.to_csv(OUTPUT_DIR / "imei_missing_from_intune_cleaned.csv", index=False)
    matched.to_csv(OUTPUT_DIR / "imei_matched_in_intune.csv", index=False)
    excluded.to_csv(OUTPUT_DIR / "imei_excluded_from_review.csv", index=False)
    disabled_action.to_csv(OUTPUT_DIR / "imei_disabled_ad_user_matches.csv", index=False)
    action_list.to_csv(OUTPUT_DIR / "imei_action_list.csv", index=False)
    intune_no_imei_report.to_csv(OUTPUT_DIR / "intune_devices_without_imei.csv", index=False)
    dept_summary_raw.to_csv(OUTPUT_DIR / "imei_department_summary_raw.csv", index=False)
    dept_summary_cleaned.to_csv(OUTPUT_DIR / "imei_department_summary_cleaned.csv", index=False)
    dept_summary_cleaned.to_csv(OUTPUT_DIR / "imei_department_summary.csv", index=False)
    summary.to_csv(OUTPUT_DIR / "imei_summary.csv", index=False)

    print("Creating master workbook...")
    write_master_workbook(
        report,
        missing_raw,
        missing_tagged,
        matched,
        excluded,
        disabled_action,
        action_list,
        intune_no_imei_report,
        dept_summary_raw,
        dept_summary_cleaned,
        classification_summary,
        top_users,
        top_departments,
        validation_samples,
        summary,
    )

    print("Creating review workbook...")
    write_review_workbook(action_list, disabled_action, excluded, classification_summary)

    print("Creating department workbook...")
    write_department_workbook(action_list, disabled_action)

    print("\nAudit Complete\n")
    print(summary.to_string(index=False))
    print(f"\nMaster workbook: {OUTPUT_DIR / MASTER_WORKBOOK}")
    print(f"Review workbook: {OUTPUT_DIR / REVIEW_WORKBOOK}")
    print(f"Department workbook: {OUTPUT_DIR / DEPT_WORKBOOK}")


if __name__ == "__main__":
    main()